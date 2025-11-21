import os
import logging
import pandas as pd
from openpyxl import load_workbook
from perfilesOcupacionales.perfilExcepcion import PerfilOcupacionalNoEncontrado
from perfilesOcupacionales.gestorDePerfilesOcupacionales import (
    extraer_nombre_ficha, 
    cargar_mapeo_perfiles, 
    buscar_perfil_ocupacional, 
    obtener_nombre_ficha
)
from funciones_excel.conversion_excel import convertir_xls_a_xlsx
from funciones_excel.extraccion_datos_excel import extraer_info_antes_conversion

# Mapeo de tipos de documento
TIPOS_DOCUMENTO = {
    "CC": "1",
    "TI": "2", 
    "CE": "3",
    "Otro Documento de Identidad": "5",
    "PEP": "8",
    "PPT": "9"
}

def preparar_excel(ruta_excel):
    """
    Prepara el archivo Excel para procesamiento
    Convierte .xls a .xlsx automáticamente
    """
    programa_sin_perfil = None
    nombre_programa = None
    perfil_encontrado = None
    
    try:
        # ===== VALIDAR ARCHIVO =====
        if not os.path.exists(ruta_excel):
            raise FileNotFoundError(f"Archivo no encontrado: {ruta_excel}")
        
        logging.info(f"Abriendo archivo: {ruta_excel}")
        
        # ===== EXTRAER INFO ANTES DE CONVERTIR =====
        if ruta_excel.endswith('.xls'):
            print(f"Archivo .xls detectado - extrayendo información...")
            nombre_ficha, nombre_programa = extraer_info_antes_conversion(ruta_excel)
            
            if not nombre_programa:
                error_msg = "✗ No se pudo obtener el nombre del programa del archivo .xls"
                logging.error(error_msg)
                print(error_msg)
                raise Exception(error_msg)
            
            # BUSCAR PERFIL ANTES DE CONVERTIR
            mapeo_perfiles = cargar_mapeo_perfiles()
            
            if mapeo_perfiles:
                logging.info(f"Mapeo cargado con {len(mapeo_perfiles)} programas")
                perfil_encontrado = buscar_perfil_ocupacional(nombre_programa, mapeo_perfiles)
                
                print(f"✓ Perfil encontrado: {perfil_encontrado}")
                
                # VALIDACIÓN ESTRICTA ANTES DE CONVERTIR
                if not perfil_encontrado or str(perfil_encontrado).strip() == '':
                    programa_sin_perfil = nombre_programa
                    
                    error_msg = f"✗ Perfil no encontrado para: {nombre_programa}"
                    logging.error(error_msg)
                    print(error_msg)
                    
                    # LANZAR EXCEPCIÓN ANTES DE CONVERTIR
                    raise PerfilOcupacionalNoEncontrado(nombre_programa)
                
                print(f"✓ Perfil válido: {perfil_encontrado}")
                logging.info(f"Perfil encontrado: {perfil_encontrado}")
            else:
                error_msg = "✗ No se pudo cargar el mapeo de perfiles"
                logging.error(error_msg)
                print(error_msg)
                raise Exception(error_msg)
            
            # Si llegamos aquí, TODO está validado - AHORA SÍ convertir
            print(f"✓ Procediendo con conversión...")
            ruta_excel = convertir_xls_a_xlsx(ruta_excel)
        
        # ===== CONFIGURAR PANDAS =====
        pd.set_option('display.float_format', lambda x: '%.0f' % x)
        
        # ===== LEER CON PANDAS =====
        df = pd.read_excel(
            ruta_excel, 
            header=4,
            dtype={'Número de Documento': str, 'Celular': str}, 
            engine='openpyxl'
        )
        
        # ===== CARGAR CON OPENPYXL =====
        wb = load_workbook(ruta_excel)
        sheet = wb.active
        
        # ===== MAPEAR COLUMNAS =====
        expected_columns = [
            'Tipo de Documento', 'Número de Documento', 'Nombre', 
            'Apellidos', 'Celular', 'Correo Electrónico', 
            'Estado', 'Perfil Ocupacional'
        ]
        
        header_row = 4
        column_indices = {}
        
        for col_idx, cell in enumerate(list(sheet.rows)[header_row], start=0):
            cell_value = cell.value
            for expected_column in expected_columns:
                if cell_value and expected_column in str(cell_value):
                    column_indices[expected_column] = col_idx
                    break
        
        logging.info(f"Columnas encontradas: {list(column_indices.keys())}")
        
        # ===== PROCESAR PERFIL OCUPACIONAL =====
        col_perfil_ocupacional = 'Perfil Ocupacional'
        necesita_guardar = False
        
        # Si NO teníamos perfil (archivo .xlsx desde el inicio), buscarlo ahora
        if not perfil_encontrado:
            mapeo_perfiles = cargar_mapeo_perfiles()
            
            if mapeo_perfiles:
                logging.info(f"Mapeo cargado con {len(mapeo_perfiles)} programas")
                print(f"✓ Mapeo de perfiles cargado")
                
                # Buscar ficha en archivo .xlsx
                ficha_caracterizacion = None
                for row in list(sheet.rows)[:5]:
                    for cell in row:
                        if cell.value and 'Ficha de Caracterización' in str(cell.value):
                            col_idx = cell.column - 1
                            if col_idx + 1 < len(row):
                                ficha_caracterizacion = row[col_idx + 1].value
                            break
                    if ficha_caracterizacion:
                        break
                
                if ficha_caracterizacion:
                    nombre_programa = extraer_nombre_ficha(str(ficha_caracterizacion))
                    
                    if nombre_programa:
                        perfil_encontrado = buscar_perfil_ocupacional(nombre_programa, mapeo_perfiles)
                        
                        if not perfil_encontrado or str(perfil_encontrado).strip() == '':
                            programa_sin_perfil = nombre_programa
                            logging.error(f"Perfil no encontrado: {nombre_programa}")
                            print(f"✗ Perfil no encontrado para: {nombre_programa}")
                            wb.close()
                            raise PerfilOcupacionalNoEncontrado(nombre_programa)
        
        if not perfil_encontrado:
            error_msg = "✗ No se pudo obtener perfil ocupacional"
            logging.error(error_msg)
            print(error_msg)
            wb.close()
            raise Exception(error_msg)
        
        # Si llegamos aquí, tenemos un perfil válido
        logging.info(f"✓ Perfil confirmado: {perfil_encontrado}")
        print(f"✓ Perfil: {perfil_encontrado}")
        
        # Asignar perfil al DataFrame
        if col_perfil_ocupacional not in df.columns:
            df[col_perfil_ocupacional] = ''
        df[col_perfil_ocupacional] = perfil_encontrado
        necesita_guardar = True
        
        # ===== VERIFICAR/CREAR COLUMNA EN EXCEL =====
        if col_perfil_ocupacional not in column_indices:
            logging.info(f"Creando columna '{col_perfil_ocupacional}'...")
            print(f"ℹ Creando columna '{col_perfil_ocupacional}'")
            
            next_col_index = max(column_indices.values()) + 1 if column_indices else len(column_indices)
            column_indices[col_perfil_ocupacional] = next_col_index
            
            sheet.cell(row=header_row + 1, column=next_col_index + 1, value=col_perfil_ocupacional)
            necesita_guardar = True
            
        elif col_perfil_ocupacional not in df.columns:
            df[col_perfil_ocupacional] = ''
        
        # ===== ESCRIBIR PERFIL EN EXCEL =====
        if perfil_encontrado and col_perfil_ocupacional in column_indices:
            col_perfil_idx = column_indices[col_perfil_ocupacional]
            doc_col_idx = column_indices.get('Número de Documento')
            
            filas_escritas = 0
            if doc_col_idx is not None:
                for row_idx in range(header_row + 2, sheet.max_row + 1):
                    doc_cell = sheet.cell(row=row_idx, column=doc_col_idx + 1)
                    if doc_cell.value and str(doc_cell.value).strip():
                        sheet.cell(row=row_idx, column=col_perfil_idx + 1, value=perfil_encontrado)
                        filas_escritas += 1
            
            print(f"✓ Perfil asignado a {filas_escritas} filas")
            logging.info(f"Perfil '{perfil_encontrado}' asignado a {filas_escritas} filas")
        
        # ===== GUARDAR =====
        if necesita_guardar:
            try:
                wb.save(ruta_excel)
                print(f"✓ Archivo guardado: {os.path.basename(ruta_excel)}")
                logging.info(f"Archivo guardado: {ruta_excel}")
                
            except Exception as e:
                logging.error(f"ERROR AL GUARDAR: {e}")
                print(f"✗ Error guardando: {e}")
                wb.close()
                raise
        
        # ===== LIMPIAR DATOS =====
        df = df.dropna(subset=['Número de Documento']).copy()
        
        logging.info(f"Total de registros válidos: {len(df)}")
        print(f"Total de registros: {len(df)}")
        
        # ===== VALIDACIÓN FINAL =====
        if df[col_perfil_ocupacional].isna().all() or (df[col_perfil_ocupacional] == '').all():
            error_msg = f"✗ ERROR: Perfil no se asignó correctamente"
            logging.error(error_msg)
            print(error_msg)
            wb.close()
            raise Exception(error_msg)
        
        print(f"✓ Validación final exitosa")
        
        # ===== RETORNAR =====
        return df, wb, sheet, sheet, column_indices, header_row, programa_sin_perfil, ruta_excel
        
    except PerfilOcupacionalNoEncontrado as e:
        print(f"🚨 Excepción capturada: {e}")
        logging.error(f"PerfilOcupacionalNoEncontrado: {e.nombre_programa}")
        raise
        
    except Exception as e:
        logging.error(f"Error configurando Excel: {e}")
        import traceback
        traceback.print_exc()
        raise