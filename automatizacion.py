# === LIBRERÍAS ESTÁNDAR DE PYTHON ===
import logging
import os
import sys
import threading
import time
# === LIBRERÍAS DE TERCEROS ===
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill, Font, Alignment
# === MÓDULOS LOCALES - Funciones de formularios ===
from funciones_formularios.campo_estrato import llenar_estrato
from funciones_formularios.campo_sueldo import llenar_formulario_sueldo
from funciones_formularios.campo_telefono_correo import llenar_formulario_telefono_correo
from funciones_formularios.experincia_laboral_campos import experiencia_laboral
from funciones_formularios.form_campo_estado_civil import llenar_formulario_estado_civil
from funciones_formularios.form_campo_perfil_ocupacional import llenar_input_perfil_ocupacional
from funciones_formularios.form_campos_nacimiento import llenar_formulario_ubicaciones_nacimiento
from funciones_formularios.form_campos_ubicacion_identificacion import llenar_formulario_ubicaciones
from funciones_formularios.form_datos_residencia import llenar_formulario_ubicacion_residencia
from funciones_formularios.login import login
from funciones_formularios.meses_busqueda import verificar_meses_busqueda
from funciones_excel.preparar_excel import preparar_excel
from funciones_formularios.pre_inscripcion import llenar_datos_antes_de_inscripcion
from funciones_formularios.verificacion import (
    verificar_estudiante,
    verificar_estudiante_con_CC_primero
)
# === MÓDULOS LOCALES - Otros ===
from funciones_loggs.loggs_funciones import loggs
from perfilesOcupacionales.perfilExcepcion import PerfilOcupacionalNoEncontrado
from URLS.urls import URL_FORMULARIO, URL_VERIFICACION
from debug_exe import close_logger, get_log_path

#Funcion que preapara los logs en un archivo y maneja su durabilidad en la aplicación
loggs()

# Cargar variables de entorno
load_dotenv()

# Mapeo de tipos de documento
TIPOS_DOCUMENTO = {
    "CC": "1",
    "TI": "2", 
    "CE": "3",
    "Otro Documento de Identidad": "5",
    "PEP": "8",
    "PPT": "9"
}

# --- Variables Globales ---
# Se definirán dentro de la función main para que sean accesibles en todo el script
RUTA_EXCEL = ""
# Estilos para .xlsx
fill_procesando = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro
fill_procesado = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # Verde claro
fill_ya_existe = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")   # Amarillo claro
fill_error = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")       # Rojo

class QueueStream:
    """Un objeto tipo archivo que escribe en una cola de progreso."""
    def __init__(self, queue):
        self.queue = queue

    def write(self, text):
        # Envía el texto a la cola para que la GUI lo muestre.
        if self.queue:
            self.queue.put(("log", text))

    def flush(self):
        # Necesario para la interfaz de archivo, pero no hace nada aquí.
        pass

def debug_log(mensaje, progress_queue=None):
    """Helper para loggear tanto en consola como en GUI"""
    print(mensaje)
    logging.info(mensaje)
    if progress_queue:
        progress_queue.put(("log", f"🔍 {mensaje}\n"))

def main(ruta_excel_param, progress_queue=None, username=None, password=None, stop_event=threading.Event(), pause_event=None):
    from debug_exe import init_logger, log, close_logger
    chrome_options = Options()

    # Descomentar la siguiente línea para modo headless (sin interfaz gráfica), para visualizar el funcionamiento del aplicativo
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")


    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    wait = WebDriverWait(driver, 10)  # Espera explícita de 10 segundos
    wait_rapido = WebDriverWait(driver, 3) # Espera mas corta


    init_logger()
    log(f"Ruta Excel: {ruta_excel_param}")
    # ===== VALIDAR EVENTOS =====
    if stop_event is None:
        stop_event = threading.Event()
    if pause_event is None:
        pause_event = threading.Event()
        pause_event.set()  # Por defecto no pausado
    
    # Hacemos globales las variables que se usarán en todo el script
    global RUTA_EXCEL, df, wb, sheet, read_sheet, column_indices, header_row, programa_sin_perfil
    RUTA_EXCEL = ruta_excel_param
    original_stdout = sys.stdout
    
    if progress_queue:
        sys.stdout = QueueStream(progress_queue)

    try:
        try:
            resultado = preparar_excel(RUTA_EXCEL)
            df, wb, sheet, read_sheet, column_indices, header_row, programa_sin_perfil, RUTA_EXCEL = resultado
            debug_log("preparar_excel completado", progress_queue)
            
        except PerfilOcupacionalNoEncontrado as e:
            nombre_programa = e.nombre_programa
            logging.warning(f"Perfil no encontrado para: {nombre_programa}")
            
            if progress_queue:
                progress_queue.put(("log", f"\n{'='*50}\n"))
                progress_queue.put(("log", f" PERFIL OCUPACIONAL NO ENCONTRADO\n"))
                progress_queue.put(("log", f"{'='*50}\n"))
                progress_queue.put(("log", f" Programa: {nombre_programa}\n"))
                progress_queue.put(("log", f"⏸ El proceso no puede continuar sin un perfil válido\n"))
                progress_queue.put(("log", f" Por favor, configura el perfil ocupacional\n\n"))
                
                # Solicitar perfil (BLOQUEARÁ hasta que el usuario responda)
                progress_queue.put(("solicitar_perfil", nombre_programa))
            
            # SALIR INMEDIATAMENTE - No continuar procesando
            close_logger()
            driver.quit()
            return
        
        # Si llegamos aquí, preparar_excel fue exitoso
        if RUTA_EXCEL != ruta_excel_param:
            print(f" Ruta actualizada a: {RUTA_EXCEL}")
            logging.info(f"Ruta de trabajo actualizada: {RUTA_EXCEL}")
        
        # ===== VERIFICAR QUE PERFIL ESTÁ EN DF =====
        if 'Perfil Ocupacional' not in df.columns or df['Perfil Ocupacional'].isna().all():
            error_msg = "✗ ERROR: No se pudo cargar el perfil ocupacional en el DataFrame"
            logging.error(error_msg)
            print(error_msg)
            
            if progress_queue:
                progress_queue.put(("log", f"\n{error_msg}\n"))
                progress_queue.put(("finish", None))
            
            close_logger()
            driver.quit()
            return
        
    except (FileNotFoundError, Exception) as e:
        logging.error(f"Error fatal al preparar el archivo Excel: {e}")
        print(f"✗ Error fatal: {e}")
        
        if progress_queue:
            progress_queue.put(("log", f"✗ Error fatal: {e}\n"))
            progress_queue.put(("finish", None))
        
        close_logger()
        driver.quit()
        return

    try:
        # ===== VERIFICAR DETENCIÓN ANTES DE LOGIN =====
        if stop_event.is_set():
            print("🛑 Proceso detenido antes de iniciar login")
            return
        
        # Realizar login
        if not login(driver, username, password):
            logging.error("No se pudo completar el login. Abortando proceso.")
            return

        # Establecer columnas
        COLUMNA_TIPO_DOC = 'Tipo de Documento'
        COLUMNA_NUM_DOC = 'Número de Documento'
        COLUMNA_NOMBRES = 'Nombre'
        COLUMNA_APELLIDOS = 'Apellidos'
        COLUMNA_CELULAR = 'Celular'
        COLUMNA_CORREO = 'Correo Electrónico'
        COLUMNA_ESTADO = 'Estado'
        COLUMNA_PERFIL = 'Perfil Ocupacional'
        
        # Contadores
        contador_procesados_exitosamente = 0
        contador_ya_existentes = 0
        contador_errores = 0
        contador_saltados = 0
        
        total_registros = len(df)
        
        # =====  BUCLE PRINCIPAL CON CONTROL DE PAUSA/DETENCIÓN =====
        for i, fila in df.iterrows():
            excel_row = i + header_row + 1
            
            # ===== VERIFICAR DETENCIÓN =====
            # Verificar si se debe detener
            if stop_event.is_set():
                progress_queue.put(("log", f"Proceso detenido por el usuario.\n"))
                return
            
            if pause_event:
                pause_event.wait() 

            # ===== VERIFICAR PAUSA =====
            if not pause_event.is_set():
                print(f"\n{'='*50}")
                print(f"⏸ PROCESO PAUSADO")
                print(f"{'='*50}")
                print(f" Pausado en registro: {i + 1}/{total_registros}")
                print(f" Esperando reanudación...")
                
                if progress_queue:
                    progress_queue.put(("log", f"\n⏸ Pausado en registro {i + 1}/{total_registros}\n"))
                
                # Esperar hasta que se reanude o se detenga
                while not pause_event.is_set():
                    if stop_event.is_set():
                        print(" Detenido durante pausa")
                        if progress_queue:
                            progress_queue.put(("log", " Proceso detenido durante pausa\n"))
                        return
                    time.sleep(0.5)  # Verificar cada 500ms
                
                print(f"\n{'='*50}")
                print(f"▶ PROCESO REANUDADO")
                print(f"{'='*50}")
                print(f" Continuando desde registro: {i + 1}/{total_registros}\n")
                
                if progress_queue:
                    progress_queue.put(("log", f"\n▶ Reanudado desde registro {i + 1}/{total_registros}\n"))
            
            # ===== ENVIAR PROGRESO =====
            if progress_queue:
                progress_queue.put(("progress", (i + 1, total_registros)))

            # Inicializar variables
            nombres = "Sin nombre"
            apellidos = "Sin apellido"
            tipo_doc = ""
            num_doc = ""
            celular = ""
            correo = ""
            estado = ""
            perfil_ocupacional = ""
            
            try:
                # ===== COLOREAR COMO PROCESANDO =====
                for col_name, col_idx in column_indices.items():
                    try:
                        if col_name in df.columns:
                            valor = fila[col_name]
                        else:
                            valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                        cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                        cell.fill = fill_procesando
                    except Exception as e:
                        print(f" Error al colorear celda {col_name}: {str(e)}")
                
                wb.save(RUTA_EXCEL)
                print(f" Excel actualizado: marcando fila {excel_row + 1} como 'procesando'")
                
                # ===== EXTRAER DATOS =====
                tipo_doc = str(fila[COLUMNA_TIPO_DOC])
                num_doc = str(fila[COLUMNA_NUM_DOC])
                nombres = str(fila[COLUMNA_NOMBRES])
                apellidos = str(fila[COLUMNA_APELLIDOS])
                celular = str(fila[COLUMNA_CELULAR])
                correo = str(fila[COLUMNA_CORREO])
                estado = str(fila[COLUMNA_ESTADO])
                perfil_ocupacional = str(fila[COLUMNA_PERFIL])
                
                logging.info(f"Procesando estudiante {i+1}/{total_registros}: {nombres} {apellidos}")
                print(f"\n{'='*50}")
                print(f" Procesando {i+1}/{total_registros}: {nombres} {apellidos}")
                print(f"{'='*50}\n")
                
                # ===== VERIFICAR DETENCIÓN ANTES DE VERIFICACIÓN =====
                if stop_event.is_set():
                    print("🛑 Detención solicitada antes de verificar estudiante")
                    break
                
                # ===== VERIFICAR SI EXISTE =====
                existe = verificar_estudiante_con_CC_primero(tipo_doc, num_doc, nombres, apellidos, driver, wait, wait_rapido)
                
                if existe is None:
                    logging.warning(f"Saltando estudiante debido a error en verificación: {nombres} {apellidos}")
                    print(f"Saltando estudiante debido a error en verificación")
                    
                    for col_name, col_idx in column_indices.items():
                        try:
                            if col_name in df.columns:
                                valor = fila[col_name]
                            else:
                                valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                            cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                            cell.fill = fill_error
                        except:
                            pass
                    wb.save(RUTA_EXCEL)
                    contador_saltados += 1
                    continue
                
                if existe:
                    logging.info(f"El estudiante {nombres} {apellidos} ya existe en el sistema")
                    print(f"✓ Estudiante ya existe en el sistema")
                    
                    for col_name, col_idx in column_indices.items():
                        try:
                            if col_name in df.columns:
                                valor = fila[col_name]
                            else:
                                valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                            cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                            cell.fill = fill_ya_existe
                        except:
                            pass
                    wb.save(RUTA_EXCEL)
                    contador_ya_existentes += 1
                    continue
                
                # ===== VERIFICAR DETENCIÓN ANTES DE REGISTRO =====
                if stop_event.is_set():
                    print("🛑 Detención solicitada antes de registrar estudiante")
                    break
                
                logging.info(f"El estudiante {nombres} {apellidos} no existe. Procediendo con el registro.")
                print(f" Estudiante no existe. Procediendo con registro...")
                
                # ===== PROCESO DE REGISTRO (con verificaciones intercaladas) =====
                if URL_VERIFICACION in driver.current_url:
                    print(" Formulario de pre-inscripción detectado")
                    
                    if stop_event.is_set():
                        break
                    
                    if llenar_datos_antes_de_inscripcion(nombres, apellidos, driver):
                        print("✓ Pre-inscripción completada")
                        
                        wait.until(EC.invisibility_of_element_located((By.ID, "content-load")))
                        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "well")))

                        if stop_event.is_set():
                            break
                        
                        ya_registrado = verificar_meses_busqueda(driver)
                        
                        if ya_registrado:
                            logging.info(f"El estudiante {nombres} {apellidos} ya está registrado")
                            print(f"✓ Estudiante ya registrado (meses de búsqueda > 1)")
                            
                            for col_name, col_idx in column_indices.items():
                                try:
                                    if col_name in df.columns:
                                        valor = fila[col_name]
                                    else:
                                        valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                    cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                    cell.fill = fill_ya_existe
                                except:
                                    pass
                            
                            wb.save(RUTA_EXCEL)
                            contador_ya_existentes += 1
                            continue
                        
                        if stop_event.is_set():
                            break
                        
                        # ===== LLENAR FORMULARIO CON VERIFICACIONES =====
                        if driver.current_url == URL_FORMULARIO or "formulario" in driver.current_url.lower():
                            print(" Llenando formulario completo...")
                            
                            # Cada función de llenado podría verificar stop_event internamente
                            llenar_formulario_ubicaciones(driver)
                            if stop_event.is_set(): break
                            
                            llenar_formulario_ubicaciones_nacimiento(driver, wait)
                            if stop_event.is_set(): break
                            
                            llenar_formulario_ubicacion_residencia(driver)
                            if stop_event.is_set(): break
                            
                            llenar_formulario_estado_civil(driver)
                            if stop_event.is_set(): break
                            
                            llenar_formulario_sueldo(driver)
                            if stop_event.is_set(): break
                            
                            llenar_estrato(driver)
                            if stop_event.is_set(): break
                            
                            llenar_formulario_telefono_correo(celular, correo, driver)
                            if stop_event.is_set(): break
                            
                            llenar_input_perfil_ocupacional(estado, driver)
                            if stop_event.is_set(): break
                            
                            print("✓ Formulario con datos básicos llenado correctamente")
                            
                            # Guardar información
                            botonGuardar = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.ID, 'submitNewOft'))
                            )
                            botonGuardar.click()
                            print("✓ Click en botón Guardar")
                            logging.info("Se hizo click en el boton de Guardar")
                            
                            if stop_event.is_set():
                                break
                            
                            resultado_experiencia_laboral = experiencia_laboral(driver, perfil_ocupacional)
                            
                            if resultado_experiencia_laboral == False:
                                for col_name, col_idx in column_indices.items():
                                    try:
                                        if col_name in df.columns:
                                            valor = fila[col_name]
                                        else:
                                            valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                        cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                        cell.fill = fill_error
                                    except:
                                        pass
                                wb.save(RUTA_EXCEL)
                                print(f"✗ Error en experiencia laboral")
                                contador_errores += 1
                                driver.get(URL_VERIFICACION)
                                time.sleep(2)
                                continue
                            
                            # ===== ÉXITO =====
                            for col_name, col_idx in column_indices.items():
                                try:
                                    if col_name in df.columns:
                                        valor = fila[col_name]
                                    else:
                                        valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                    cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                    cell.fill = fill_procesado
                                except:
                                    pass
                            wb.save(RUTA_EXCEL)
                            print(f"✓ Registro completado exitosamente")
                            contador_procesados_exitosamente += 1

                        else:
                            print(f" No se detectó formulario completo. URL: {driver.current_url}")
                            for col_name, col_idx in column_indices.items():
                                try:
                                    valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                    cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                    cell.fill = fill_error
                                except:
                                    pass
                            wb.save(RUTA_EXCEL)
                            contador_errores += 1
                    else:
                        print("✗ No se pudo completar la pre-inscripción")
                        for col_name, col_idx in column_indices.items():
                            try:
                                if col_name in df.columns:
                                    valor = fila[col_name]
                                else:
                                    valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                cell.fill = fill_error
                            except:
                                pass
                        wb.save(RUTA_EXCEL)
                        contador_errores += 1
                else:
                    print(f" No se redirigió al formulario. URL: {driver.current_url}")
                    driver.get(URL_VERIFICACION)
                    print(" Reintentando verificación...")
                    existe = verificar_estudiante(tipo_doc, num_doc, nombres, apellidos, driver, wait, wait_rapido)
                    if not existe:
                        print(" Reintentando llenar datos...")
                        if llenar_datos_antes_de_inscripcion(nombres, apellidos, driver, wait):
                            for col_name, col_idx in column_indices.items():
                                try:
                                    valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                    cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                    cell.fill = fill_procesado
                                except:
                                    pass
                            contador_procesados_exitosamente += 1
                        else:
                            for col_name, col_idx in column_indices.items():
                                try:
                                    valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                                    cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                                    cell.fill = fill_error
                                except:
                                    pass
                            contador_errores += 1
                        wb.save(RUTA_EXCEL)
                    
            except Exception as e:
                logging.error(f"Error procesando estudiante {nombres} {apellidos}: {str(e)}")
                print(f"✗ Error procesando estudiante: {str(e)}")
                
                for col_name, col_idx in column_indices.items():
                    try:
                        if col_name in df.columns:
                            valor = fila[col_name]
                        else:
                            valor = sheet.cell(row=excel_row + 1, column=col_idx + 1).value  # Para openpyxl las filas empiezan en 1
                            cell = sheet.cell(row=excel_row + 1, column=col_idx + 1, value=valor)
                            cell.fill = fill_error
                    except:
                        pass
                
                try:
                    wb.save(RUTA_EXCEL)
                    print(f" Excel actualizado: marcando como error")
                except Exception as save_error:
                    print(f" Error al guardar Excel: {str(save_error)}")
                
                contador_errores += 1
                
                try:
                    driver.get(URL_VERIFICACION)
                    time.sleep(2)
                except Exception as nav_error:
                    print(f" Error al navegar: {str(nav_error)}")
        
        # ===== RESUMEN FINAL =====
        try:
            fila_resumen = total_registros + header_row + 3 + 1  # +1 porque openpyxl empieza en 1
            # Estilos
            fill_titulo = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            font_titulo = Font(bold=True, color="FFFFFF")
            font_resumen = Font(bold=True)
            # Título
            cell_titulo = sheet.cell(row=fila_resumen, column=1, value="RESUMEN DE PROCESAMIENTO")
            cell_titulo.fill = fill_titulo
            cell_titulo.font = font_titulo
            # Datos
            sheet.cell(row=fila_resumen + 1, column=1, value="Aprendices ingresados exitosamente:").font = font_resumen
            sheet.cell(row=fila_resumen + 1, column=2, value=contador_procesados_exitosamente).font = font_resumen
            
            sheet.cell(row=fila_resumen + 2, column=1, value="Ya existentes:").font = font_resumen
            sheet.cell(row=fila_resumen + 2, column=2, value=contador_ya_existentes).font = font_resumen
            
            sheet.cell(row=fila_resumen + 3, column=1, value="Errores:").font = font_resumen
            sheet.cell(row=fila_resumen + 3, column=2, value=contador_errores).font = font_resumen
            
            sheet.cell(row=fila_resumen + 4, column=1, value="Saltados:").font = font_resumen
            sheet.cell(row=fila_resumen + 4, column=2, value=contador_saltados).font = font_resumen
            
            sheet.cell(row=fila_resumen + 5, column=1, value="Total procesados:").font = font_resumen
            sheet.cell(row=fila_resumen + 5, column=2, value=total_registros).font = font_resumen

            wb.save(RUTA_EXCEL)
            print("\n Resumen guardado en Excel")
        except Exception as e:
            print(f" Error al escribir resumen: {str(e)}")
            logging.error(f"Error al escribir resumen: {str(e)}")
        
        # ===== MENSAJE FINAL =====
        if stop_event.is_set():
            logging.info(" Proceso detenido por el usuario")
            print("\n Proceso detenido por el usuario")
        else:
            logging.info("Proceso completado exitosamente")
            print("\n Proceso completado exitosamente")
            
    except Exception as e:
        logging.error(f"Error general en el proceso: {str(e)}")
        print(f" Error general: {str(e)}")
        
    finally:
        sys.stdout = original_stdout
        try:
            driver.quit()
            logging.info("Navegador cerrado")
            print("Navegador cerrado")
        except Exception as e:
            logging.error(f"Error al cerrar navegador: {e}")
        close_logger()
        
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        ruta_archivo_excel = sys.argv[1]
        main(ruta_archivo_excel)
